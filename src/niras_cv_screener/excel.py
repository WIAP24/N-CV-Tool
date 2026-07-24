from __future__ import annotations

import json
from copy import copy
from pathlib import Path
from typing import Any, Dict, List

from openpyxl.utils import get_column_letter

from .criteria import criteria_to_rows


def write_workbook(
    out_path: Path,
    criteria_json: Dict[str, Any],
    results: List[Dict[str, Any]],
    errors: List[Dict[str, str]],
    extraction_warnings: List[Dict[str, str]],
    run_settings: Dict[str, Any],
    calibration_rows: List[Dict[str, Any]] | None = None,
    review_queue: List[Dict[str, Any]] | None = None,
    model_evaluations: List[Dict[str, Any]] | None = None,
    model_evaluation_details: List[Dict[str, Any]] | None = None,
    cost_records: List[Dict[str, Any]] | None = None,
    cost_summary: Dict[str, Any] | None = None,
    cost_preview: Dict[str, Any] | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    calibration_rows = calibration_rows or []
    review_queue = review_queue or []
    model_evaluations = model_evaluations or []
    model_evaluation_details = model_evaluation_details or []
    cost_records = cost_records or []
    cost_summary = cost_summary or {}
    cost_preview = cost_preview or {}

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="EADCF8")
    pass_fill = PatternFill("solid", fgColor="D9EAD3")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")
    reserve_fill = PatternFill("solid", fgColor="FFF2CC")

    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = [
        "Candidate Name",
        "File Name",
        "Recommendation",
        "Meets Mandatory",
        "Mandatory Pass Count",
        "Mandatory Total",
        "Mandatory Average",
        "Preferred Average",
        "Weighted Score",
        "Critical Gaps",
        "Quality Flags",
        "Compliance Flags",
        "Review Items",
        "Model Comparison",
    ]
    write_header(ws_summary, summary_headers, header_fill)
    review_counts = count_by_file(review_queue, "candidate_file")
    for row_num, result in enumerate(results, start=2):
        summary = result.get("summary", {})
        comparison = result.get("model_comparison") or {}
        comparison_label = ""
        if comparison:
            comparison_label = "Recommendation changed" if comparison.get("recommendation_changed") else f"Max score delta: {comparison.get('max_abs_score_delta', 0)}"
        values = [
            result.get("candidate_name", "Unknown"),
            result.get("candidate_file", result.get("candidate_id", "")),
            summary.get("recommendation", ""),
            summary.get("meets_all_mandatory", False),
            summary.get("mandatory_pass_count", 0),
            summary.get("mandatory_total", 0),
            summary.get("mandatory_average", 0),
            summary.get("preferred_average", 0),
            summary.get("weighted_score", 0),
            "\n".join(summary.get("critical_gaps", [])),
            "\n".join(summary.get("quality_flags", [])),
            len(result.get("compliance_flags", [])),
            review_counts.get(result.get("candidate_file", ""), 0),
            comparison_label,
        ]
        for col_num, value in enumerate(values, start=1):
            cell = ws_summary.cell(row=row_num, column=col_num, value=safe_value(value))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        rec = summary.get("recommendation", "")
        fill = pass_fill if rec == "Interview" else reserve_fill if rec == "Reserve" else fail_fill
        ws_summary.cell(row=row_num, column=3).fill = fill
    set_widths(ws_summary, [26, 40, 16, 18, 20, 16, 18, 18, 16, 60, 60, 16, 14, 28])
    ws_summary.freeze_panes = "A2"

    ws_matrix = wb.create_sheet("Screening Matrix")
    criteria_rows = criteria_to_rows(criteria_json)
    candidate_ids = [result.get("candidate_id", result.get("candidate_file", "")) for result in results]
    display_names = [result.get("candidate_name") or result.get("candidate_file", "") for result in results]
    matrix_headers = ["Section", "Criterion ID", "Mandatory", "Pass Score", "Weight", "Criterion"] + display_names
    write_header(ws_matrix, matrix_headers, header_fill)
    by_candidate = {
        cid: {item["criterion_id"]: item for item in result.get("assessments", [])}
        for cid, result in zip(candidate_ids, results)
    }
    for row_num, criterion in enumerate(criteria_rows, start=2):
        values = [criterion["section"], criterion["id"], criterion["mandatory"], criterion["pass_score"], criterion["weight"], criterion["text"]]
        for col_num, value in enumerate(values, start=1):
            ws_matrix.cell(row=row_num, column=col_num, value=safe_value(value))
        for offset, cid in enumerate(candidate_ids, start=7):
            assessment = by_candidate.get(cid, {}).get(criterion["id"])
            score = assessment.get("score") if assessment else None
            cell = ws_matrix.cell(row=row_num, column=offset, value=score)
            if score is not None:
                cell.fill = pass_fill if int(score) >= int(criterion["pass_score"]) else fail_fill
    set_widths(ws_matrix, [24, 16, 12, 12, 10, 70] + [24 for _ in display_names])
    ws_matrix.freeze_panes = "G2"

    ws_evidence = wb.create_sheet("Evidence")
    evidence_headers = ["Candidate Name", "File Name", "Section", "Criterion ID", "Criterion", "Mandatory", "Score", "Pass Score", "Evidence", "Source Ref", "Confidence", "Notes"]
    write_header(ws_evidence, evidence_headers, header_fill)
    row_num = 2
    for result in results:
        for item in result.get("assessments", []):
            values = [
                result.get("candidate_name", "Unknown"), result.get("candidate_file", ""), item.get("section", ""), item.get("criterion_id", ""),
                item.get("criterion_text", ""), item.get("mandatory", False), item.get("score", 0), item.get("pass_score", 3), item.get("evidence", ""),
                item.get("source_ref", ""), item.get("confidence", ""), item.get("notes", ""),
            ]
            for col_num, value in enumerate(values, start=1):
                cell = ws_evidence.cell(row=row_num, column=col_num, value=safe_value(value))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_num += 1
    set_widths(ws_evidence, [26, 40, 24, 16, 70, 12, 10, 12, 80, 18, 14, 60])
    ws_evidence.freeze_panes = "A2"

    write_rows_sheet(wb, "Review Queue", review_queue, ["severity", "candidate_name", "candidate_file", "area", "criterion_id", "reason", "suggested_action", "reviewer_decision", "reviewer_notes"], header_fill, [12, 26, 40, 24, 16, 70, 70, 24, 60])
    write_rows_sheet(wb, "Calibration", calibration_rows, ["criterion_id", "criterion_text", "section", "mandatory", "pass_score", "candidate_count", "average_score", "min_score", "max_score", "score_spread", "score_stdev", "pass_rate", "flags", "highest_candidate", "highest_evidence", "lowest_candidate", "lowest_evidence"], header_fill, [16, 70, 24, 12, 12, 16, 14, 10, 10, 12, 12, 10, 40, 26, 60, 26, 60])
    write_rows_sheet(wb, "Model Evaluation", model_evaluations, ["candidate_name", "candidate_file", "primary_model", "comparison_model", "primary_recommendation", "comparison_recommendation", "recommendation_changed", "primary_weighted_score", "comparison_weighted_score", "weighted_score_delta", "criteria_checked", "criteria_with_notable_differences", "pass_fail_flips", "max_abs_score_delta"], header_fill, [26, 40, 22, 22, 18, 20, 20, 18, 20, 16, 16, 24, 14, 18])
    write_rows_sheet(wb, "Model Eval Details", model_evaluation_details, ["candidate_name", "candidate_file", "primary_model", "comparison_model", "criterion_id", "criterion_text", "primary_score", "comparison_score", "score_delta", "pass_fail_flip", "primary_evidence", "comparison_evidence"], header_fill, [26, 40, 22, 22, 16, 70, 14, 16, 12, 14, 70, 70])

    compliance_rows: List[Dict[str, Any]] = []
    for result in results:
        for flag in result.get("compliance_flags", []):
            row = dict(flag)
            row["candidate_name"] = result.get("candidate_name", "Unknown")
            row["candidate_file"] = result.get("candidate_file", "")
            compliance_rows.append(row)
    write_rows_sheet(wb, "Compliance", compliance_rows, ["candidate_name", "candidate_file", "category", "source_ref", "reason"], header_fill, [26, 40, 28, 18, 100])

    ws_criteria = wb.create_sheet("Criteria")
    write_header(ws_criteria, ["Section", "Criterion ID", "Mandatory", "Pass Score", "Weight", "Criterion"], header_fill)
    for row_num, criterion in enumerate(criteria_rows, start=2):
        for col_num, key in enumerate(["section", "id", "mandatory", "pass_score", "weight", "text"], start=1):
            cell = ws_criteria.cell(row=row_num, column=col_num, value=safe_value(criterion[key]))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    set_widths(ws_criteria, [24, 16, 12, 12, 10, 80])
    ws_criteria.freeze_panes = "A2"

    write_rows_sheet(wb, "Skipped files", errors, ["file", "error"], header_fill, [45, 120])
    write_rows_sheet(wb, "Extraction Warnings", extraction_warnings, ["file", "warning"], header_fill, [45, 120])

    ws_cost = wb.create_sheet("Cost Dashboard")
    row_num = 1
    write_section_title(ws_cost, row_num, "Pre-run Cost Estimate", section_fill)
    row_num += 1
    write_header(ws_cost, ["Metric", "Value"], header_fill, row=row_num)
    preview_metrics = [
        ("Pricing source", cost_preview.get("pricing_source", cost_summary.get("pricing_source", ""))),
        ("Estimate basis", cost_preview.get("estimate_basis", "")),
        ("Estimated files", cost_preview.get("estimated_files", "")),
        ("Estimated model calls", cost_preview.get("estimated_model_calls", "")),
        ("Estimated input tokens", cost_preview.get("estimated_input_tokens", "")),
        ("Estimated output tokens", cost_preview.get("estimated_output_tokens", "")),
        ("Estimated total tokens", cost_preview.get("estimated_total_tokens", "")),
        ("Estimated cost before app cache USD", cost_preview.get("estimated_uncached_cost_usd", "")),
    ]
    row_num += 1
    for metric, value in preview_metrics:
        ws_cost.cell(row=row_num, column=1, value=metric)
        ws_cost.cell(row=row_num, column=2, value=safe_value(value))
        row_num += 1

    row_num += 1
    write_section_title(ws_cost, row_num, "Selected Model Rate Card", section_fill)
    row_num += 1
    preview_headers = ["stage", "model", "model_label", "files", "estimated_calls", "estimated_input_tokens", "estimated_output_tokens", "input_usd_per_1m", "cached_input_usd_per_1m", "output_usd_per_1m", "estimated_uncached_cost_usd"]
    write_header(ws_cost, preview_headers, header_fill, row=row_num)
    row_num += 1
    for row in cost_preview.get("rows", []):
        for col_num, header in enumerate(preview_headers, start=1):
            ws_cost.cell(row=row_num, column=col_num, value=safe_value(row.get(header, "")))
        row_num += 1

    row_num += 1
    write_section_title(ws_cost, row_num, "Actual Run Cost Summary", section_fill)
    row_num += 1
    write_header(ws_cost, ["Metric", "Value"], header_fill, row=row_num)
    row_num += 1
    for key, value in cost_summary.items():
        ws_cost.cell(row=row_num, column=1, value=key)
        ws_cost.cell(row=row_num, column=2, value=safe_value(value))
        row_num += 1

    row_num += 1
    write_section_title(ws_cost, row_num, "Actual Model Calls", section_fill)
    row_num += 1
    headers = ["candidate_file", "stage", "model", "cached", "input_tokens", "cached_input_tokens", "output_tokens", "total_tokens", "token_source", "cost_usd", "uncached_estimate_usd"]
    write_header(ws_cost, headers, header_fill, row=row_num)
    row_num += 1
    for record in cost_records:
        for col_num, header in enumerate(headers, start=1):
            ws_cost.cell(row=row_num, column=col_num, value=safe_value(record.get(header, "")))
        row_num += 1
    set_widths(ws_cost, [32, 40, 22, 12, 14, 18, 14, 14, 16, 14, 20])
    ws_cost.freeze_panes = "A3"

    ws_manifest = wb.create_sheet("Run Settings")
    write_header(ws_manifest, ["Setting", "Value"], header_fill)
    for row_num, (key, value) in enumerate(run_settings.items(), start=2):
        ws_manifest.cell(row=row_num, column=1, value=key)
        ws_manifest.cell(row=row_num, column=2, value=safe_value(value))
    set_widths(ws_manifest, [32, 100])

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_rows_sheet(wb: Any, title: str, rows: List[Dict[str, Any]], headers: List[str], fill: Any, widths: List[int]) -> Any:
    sheet = wb.create_sheet(title)
    write_header(sheet, headers, fill)
    for row_num, row in enumerate(rows, start=2):
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=row_num, column=col_num, value=safe_value(row.get(header, "")))
    set_widths(sheet, widths)
    sheet.freeze_panes = "A2"
    return sheet


def write_header(sheet: Any, headers: List[str], fill: Any, row: int = 1) -> None:
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col_num, value=header)
        cell.fill = fill
        font = copy(cell.font)
        font.bold = True
        cell.font = font


def write_section_title(sheet: Any, row: int, title: str, fill: Any) -> None:
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = fill
    font = copy(cell.font)
    font.bold = True
    cell.font = font


def set_widths(sheet: Any, widths: List[int]) -> None:
    for col_num, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_num)].width = width


def count_by_file(rows: List[Dict[str, Any]], key: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for row in rows:
        value = str(row.get(key, ""))
        if value:
            counts[value] = counts.get(value, 0) + 1
    return counts


def safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value).encode("utf-8", errors="ignore").decode("utf-8")



